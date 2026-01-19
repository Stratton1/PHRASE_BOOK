"""
Setup Database Script
Creates the blank Excel structure (Master_Phrase_Library.xlsx) with proper
worksheets, column headers, and data validation rules.

Run: python 1_setup_database.py
"""

import sys
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    MASTER_DB_FILE,
    MASTER_DB_SHEET_NAME,
    SECTIONS,
    STANDARD_COLUMNS,
    CONDITION_RATINGS,
    PROPERTY_STYLES,
    PROPERTY_TYPES,
    PROPERTY_AGE_BANDS
)

# ============================================================================
# LOGGING SETUP
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def create_header_row(ws, columns):
    """
    Create formatted header row in worksheet.

    Args:
        ws: openpyxl worksheet object
        columns: list of column names
    """
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def add_data_validation(ws, column_name, col_idx, data_list, start_row=2, end_row=1000):
    """
    Add dropdown validation to a specific column.

    Args:
        ws: openpyxl worksheet object
        column_name: name of the column (for reference)
        col_idx: column index (1-based)
        data_list: list of valid values for dropdown
        start_row: first row to apply validation (default: 2, after header)
        end_row: last row to apply validation
    """
    try:
        # Create validation with quoted list
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(str(v) for v in data_list)}"',
            allow_blank=False
        )
        dv.error = f"Please select a valid {column_name}"
        dv.errorTitle = "Invalid Entry"
        dv.prompt = f"Select from {column_name} list"
        dv.promptTitle = column_name

        ws.add_data_validation(dv)

        # Apply validation to column range
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        dv.add(cell_range)

        logger.info(f"Added validation for '{column_name}' (Column {col_letter})")
    except Exception as e:
        logger.error(f"Failed to add validation for '{column_name}': {e}")


def set_column_widths(ws, columns):
    """
    Set appropriate column widths based on content.

    Args:
        ws: openpyxl worksheet object
        columns: list of column names
    """
    width_map = {
        "Section": 15,
        "Element": 25,
        "Sub_Section": 20,
        "Content": 50,
        "Condition_Rating": 12,
        "Property_Style": 18,
        "Property_Type": 15,
        "Property_Age": 18,
        "Source_File": 25
    }

    for col_idx, column_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        width = width_map.get(column_name, 15)
        ws.column_dimensions[col_letter].width = width


def create_master_sheet(wb):
    """
    Create and format the Master sheet.

    Args:
        wb: openpyxl Workbook object

    Returns:
        The Master worksheet
    """
    logger.info("Creating Master sheet...")

    # Remove default sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create Master sheet
    ws = wb.create_sheet(MASTER_DB_SHEET_NAME, 0)

    # Add headers
    create_header_row(ws, STANDARD_COLUMNS)

    # Set column widths
    set_column_widths(ws, STANDARD_COLUMNS)

    # Find column indices for validation
    col_indices = {col: idx + 1 for idx, col in enumerate(STANDARD_COLUMNS)}

    # Add data validation for specific columns
    add_data_validation(
        ws,
        "Condition_Rating",
        col_indices["Condition_Rating"],
        [str(r) for r in CONDITION_RATINGS]
    )

    add_data_validation(
        ws,
        "Property_Style",
        col_indices["Property_Style"],
        PROPERTY_STYLES
    )

    add_data_validation(
        ws,
        "Property_Type",
        col_indices["Property_Type"],
        PROPERTY_TYPES
    )

    add_data_validation(
        ws,
        "Property_Age",
        col_indices["Property_Age"],
        PROPERTY_AGE_BANDS
    )

    logger.info("Master sheet created successfully")
    return ws


def create_section_sheets(wb):
    """
    Create worksheets for each survey section.

    Args:
        wb: openpyxl Workbook object
    """
    logger.info(f"Creating {len(SECTIONS)} section sheets...")

    for section in SECTIONS:
        ws = wb.create_sheet(section)

        # Add headers
        create_header_row(ws, STANDARD_COLUMNS)

        # Set column widths
        set_column_widths(ws, STANDARD_COLUMNS)

        # Find column indices for validation
        col_indices = {col: idx + 1 for idx, col in enumerate(STANDARD_COLUMNS)}

        # Add data validation
        add_data_validation(
            ws,
            "Condition_Rating",
            col_indices["Condition_Rating"],
            [str(r) for r in CONDITION_RATINGS]
        )

        add_data_validation(
            ws,
            "Property_Style",
            col_indices["Property_Style"],
            PROPERTY_STYLES
        )

        add_data_validation(
            ws,
            "Property_Type",
            col_indices["Property_Type"],
            PROPERTY_TYPES
        )

        add_data_validation(
            ws,
            "Property_Age",
            col_indices["Property_Age"],
            PROPERTY_AGE_BANDS
        )

        logger.info(f"Created sheet: {section}")


def setup_database():
    """
    Main function to set up the database structure.
    """
    logger.info("=" * 60)
    logger.info("PHRASE LIBRARY ENGINE - DATABASE SETUP")
    logger.info("=" * 60)

    try:
        # Create workbook
        logger.info("Creating new workbook...")
        wb = Workbook()

        # Create Master sheet
        create_master_sheet(wb)

        # Create section sheets
        create_section_sheets(wb)

        # Save workbook
        output_path = Path(MASTER_DB_FILE)
        wb.save(output_path)
        logger.info(f"✓ Database file saved: {output_path.absolute()}")

        # Verify file exists
        if output_path.exists():
            file_size = output_path.stat().st_size / 1024  # KB
            logger.info(f"✓ File size: {file_size:.2f} KB")
            logger.info(f"✓ Sheets created: {', '.join(wb.sheetnames)}")
            logger.info("=" * 60)
            logger.info("DATABASE SETUP COMPLETE")
            logger.info("=" * 60)
            return True
        else:
            logger.error("Failed to create database file")
            return False

    except Exception as e:
        logger.error(f"Database setup failed: {e}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    success = setup_database()
    sys.exit(0 if success else 1)
